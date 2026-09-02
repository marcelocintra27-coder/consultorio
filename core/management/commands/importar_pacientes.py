import openpyxl
from datetime import datetime, date
from django.core.management.base import BaseCommand
from core.models import Paciente, Convenio, Evolucao


def parse_data(valor):
    if valor is None or valor == "":
        return None
    if isinstance(valor, (datetime, date)):
        return valor if isinstance(valor, date) and not isinstance(valor, datetime) else valor.date()
    if isinstance(valor, str):
        valor = valor.strip()
        for fmt in ("%d/%m/%Y", "%d/%m/%y"):
            try:
                return datetime.strptime(valor, fmt).date()
            except ValueError:
                continue
    return None


def texto(valor):
    if valor is None:
        return ""
    return str(valor).strip()


class Command(BaseCommand):
    help = "Importa pacientes e evolucoes da planilha_migracao_pacientes.xlsx"

    def add_arguments(self, parser):
        parser.add_argument("--arquivo", default="planilha_migracao_pacientes.xlsx")

    def handle(self, *args, **options):
        caminho = options["arquivo"]
        try:
            wb = openpyxl.load_workbook(caminho, data_only=True)
        except FileNotFoundError:
            self.stderr.write(self.style.ERROR(f"Arquivo nao encontrado: {caminho}"))
            return

        ws1 = wb["Cadastro_Pacientes"]
        headers = [c.value for c in ws1[1]]

        def idx(nome_coluna):
            return headers.index(nome_coluna)

        criados_pac = 0
        atualizados_pac = 0
        erros_pac = []
        convenios_cache = {}

        for row_num, row in enumerate(ws1.iter_rows(min_row=2, values_only=True), start=2):
            nome = texto(row[idx("Nome Completo")])
            if not nome:
                continue
            try:
                data_nasc = parse_data(row[idx("Data de Nascimento")])
                if not data_nasc:
                    erros_pac.append(f"Linha {row_num}: data de nascimento invalida para {nome} - pulado")
                    continue
                telefone1 = texto(row[idx("Telefone 1")])
                if not telefone1:
                    erros_pac.append(f"Linha {row_num}: telefone ausente para {nome} - pulado")
                    continue
                nome_convenio = texto(row[idx("Convenio")])
                convenio_obj = None
                if nome_convenio and nome_convenio.lower() not in ("nenhum", "n/a", "-"):
                    if nome_convenio not in convenios_cache:
                        convenio_obj, _ = Convenio.objects.get_or_create(nome=nome_convenio)
                        convenios_cache[nome_convenio] = convenio_obj
                    else:
                        convenio_obj = convenios_cache[nome_convenio]

                partes_obs = []
                for campo in ["Queixa Principal (QP)", "Historico Doenca Atual (HMA)", "Historico Doenca Pregressa (HP)", "Historico Familiar (HF)", "Observacoes Gerais"]:
                    v = texto(row[idx(campo)])
                    if v:
                        partes_obs.append(f"{campo}: {v}")

                campos_saude = ["Asma", "Tosse", "Falta de Ar", "Sinusite", "Fuma", "Cardiopatia", "Febre Reumatica", "Artrite", "Radioterapia", "Disturbio Sanguineo", "Hemorragia", "Corta/Cicatriza", "Ictericia", "Hepatite", "Cirrose", "Bebe", "Diabetes", "Convulsao", "Desmaios", "Tratamento Psiquiatrico", "Cefaleias", "Medo ou Trauma", "Doenca Grave", "Gravidez"]
                positivos = [c for c in campos_saude if texto(row[idx(c)]).lower() in ("sim", "s", "yes")]
                if positivos:
                    partes_obs.append("Condicoes de saude relatadas: " + ", ".join(positivos))

                observacoes_final = "\n".join(partes_obs)
                cpf_valor = f"SEMCPF-{nome[:20].upper().replace(' ', '')}-{data_nasc.strftime('%d%m%Y')}"

                paciente, created = Paciente.objects.update_or_create(
                    nome_completo=nome,
                    data_nascimento=data_nasc,
                    defaults=dict(
                        cpf=cpf_valor,
                        telefone=telefone1,
                        whatsapp=texto(row[idx("Telefone 2")]),
                        email="",
                        endereco=texto(row[idx("Endereco")]),
                        convenio=convenio_obj,
                        observacoes=observacoes_final,
                    ),
                )
                if created:
                    criados_pac += 1
                else:
                    atualizados_pac += 1
            except Exception as e:
                erros_pac.append(f"Linha {row_num}: erro inesperado para {nome}: {e}")

        self.stdout.write(self.style.SUCCESS(f"Pacientes: {criados_pac} criados, {atualizados_pac} atualizados."))
        for e in erros_pac:
            self.stdout.write(self.style.WARNING(e))

        if "Evolucao_Atendimentos" not in wb.sheetnames:
            self.stdout.write(self.style.WARNING("Aba Evolucao_Atendimentos nao encontrada - pulando."))
            return

        ws2 = wb["Evolucao_Atendimentos"]
        criados_evo = 0
        erros_evo = []

        for row_num, row in enumerate(ws2.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:
                continue
            nome_paciente = texto(row[0])
            data_atend = parse_data(row[1])
            descricao = texto(row[2])
            if not data_atend or not descricao:
                erros_evo.append(f"Linha {row_num}: data ou descricao ausente - pulado")
                continue
            paciente = Paciente.objects.filter(nome_completo=nome_paciente).first()
            if not paciente:
                erros_evo.append(f"Linha {row_num}: paciente {nome_paciente} nao encontrado - pulado")
                continue
            Evolucao.objects.get_or_create(paciente=paciente, data=data_atend, descricao=descricao)
            criados_evo += 1

        self.stdout.write(self.style.SUCCESS(f"Evolucoes: {criados_evo} registros importados."))
        for e in erros_evo:
            self.stdout.write(self.style.WARNING(e))
